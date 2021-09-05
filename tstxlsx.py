# Test reading a xlsx file
# 09-05-2021 - RAH - New

import os
import sys, getopt
import re
import openpyxl

progname = sys.argv[0]
argv = sys.argv[1:]

book = openpyxl.load_workbook('c:\pyprograms\FinalFormatExample_JH.xlsx')

sheet_obj = book.active

m_row = sheet_obj.max_row

# Loop will print all values
# of first column
for i in range(1, m_row + 1):
    cell_obj1 = sheet_obj.cell(row=i, column=1)
    cell_obj2 = sheet_obj.cell(row=i, column=2)
    cell_obj3 = sheet_obj.cell(row=i, column=3)
    cell_obj4 = sheet_obj.cell(row=i, column=4)
    cell_obj5 = sheet_obj.cell(row=i, column=5)
    print(cell_obj1.value, cell_obj2.value, cell_obj3.value, cell_obj4.value, cell_obj5.value)
sys.exit(0)
